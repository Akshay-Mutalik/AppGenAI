from flask import Flask, render_template, request, redirect, url_for, session, flash
import pymysql
import os
from werkzeug.utils import secure_filename
import openai
import base64
import io
from functools import wraps # Import wraps for decorator
from flask import jsonify
from flask import Flask   
from flask_cors import CORS
import re 
import json

# Import Presidio libraries for PII data removal
from presidio_analyzer import AnalyzerEngine, RecognizerRegistry
# from presidio_anonymizer import AnonymizerEngine, AnonymizerConfig
# from presidio_anonymizer.entities import ReplaceConfig
# from presidio_anonymizer import AnonymizerEngine


# Import libraries for reading other file types (keep if still needed for non-image files)
try:
    import PyPDF2
except ImportError:
    PyPDF2 = None
    print("PyPDF2 not installed. PDF file reading will not be supported.")
try:
    import docx
except ImportError:
    docx = None
    print("python-docx not installed. DOCX file reading will not be supported.")
try:
    import openpyxl
except ImportError:
    openpyxl = None
    print("openpyxl not installed. XLSX file reading will not be supported.")
try:
    from PIL import Image
except ImportError:
    Image = None
    print("Pillow not installed. Image processing might be limited.")

import base64
import io
import re
try:
    import openpyxl
except ImportError:
    openpyxl = None

app = Flask(__name__)
app.secret_key = "TechMahindra" # Change this to a secure random key

CORS(app, methods=['GET', 'POST', 'PUT', 'DELETE'])
CORS(app, resources={r"/*": {"origins": "http://10.181.57.4:8899"}})
CORS(app, supports_credentials=True)
# Set Max content limit
app.config['MAX_CONTENT_LENGTH'] = 100 * 1024 * 1024 #set to 100MB
# Configure the upload folder
app.config['UPLOAD_FOLDER'] = 'uploads' 

#headers policy

@app.after_request
def add_referrer_policy(response):
    response.headers['Referrer-Policy'] = 'no-referrer'
    return response


# Database configuration
db_config = {
    'host': 'localhost',
    'user': 'root',
    'password': 'root',
    'database': 'prompt_framework_new',
    'port': 3306
}

# Azure OpenAI configuration (REPLACE WITH YOUR ACTUAL CREDENTIALS AND DEPLOYMENT)
azure_openai_config = {
    "api_key": "6dYVlazjWM6mUGSUrHkQsnLmzlcCQQNwG76L1z5aZJwu6oBGAQQoJQQJ99BDACYeBjFXJ3w3AAABACOGCPNf",  # <--- REPLACE WITH YOUR AZURE OPENAI API KEY
    "endpoint": "https://signifyopenai.openai.azure.com/",  # <--- REPLACE WITH YOUR AZURE OPENAI ENDPOINT (e.g., https://YOUR_RESOURCE_NAME.openai.azure.com/)
    "deployment_name": "gpt-4o",  # <--- REPLACE WITH YOUR ACTUAL DEPLOYMENT NAME
    "api_version": "2025-01-01-preview",
}

config_table = []

def config_init():
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("SHOW COLUMNS FROM configuration_items")
    config_table = [row[0] for row in cursor.fetchall()]
    cursor.close()
    conn.close()
    return config_table




# Decorator to check if user is logged in
def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'logged_in' not in session or not session['logged_in']:
            flash('Please log in to access this page.', 'danger')
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function

# Decorator to check if user has admin role
def admin_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'role' not in session or session['role'] != 'admin':
            flash('You do not have permission to access this page.', 'danger')
            return redirect(url_for('index')) # Redirect to home or login
        return f(*args, **kwargs)
    return decorated_function

# Decorator to check if user has bot_user role or admin role
def chatbot_access_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'role' not in session or (session['role'] != 'bot' and session['role'] != 'admin'): # Changed 'bot_user' to 'bot'
            flash('You do not have permission to access the chatbot.', 'danger')
            return redirect(url_for('index')) # Redirect to home or login
        return f(*args, **kwargs)
    return decorated_function

# Function to connect to the database
def get_db_connection():
    try:
        connection = pymysql.connect(**db_config)
        return connection
    except pymysql.MySQLError as e:
        print(f"Error connecting to MySQL: {e}")
        flash("Database connection failed. Please check your configuration and try again.", 'error')
        return None

#Function to remove PII data using presidio
def redact_pii(text):
    """
    Redacts PII from the given text using the Presidio Analyzer.
    """
    analyzer = AnalyzerEngine()
    results = analyzer.analyze(text, language='en')  # You might need to adjust the language

    redacted_text = text
    offset = 0
    for res in sorted(results, key=lambda x: x.start):
        start = res.start + offset
        end = res.end + offset
        redacted_text = redacted_text[:start] + "[REDACTED]" + redacted_text[end:]
        offset += len("[REDACTED]") - (end - start)
    return redacted_text

# To check if the response content is Excel
def is_likely_excel(text):
    """
    Attempts to detect tabular data in text based on common Excel/TSV patterns.
    This is a heuristic and may not be perfect.
    """
    lines = text.strip().split('\n')
    if not lines:
        return False
    header_line = lines[0]
    separators = ['|', '\t']
    for sep in separators:
        header_cells = len(header_line.split(sep))
        if header_cells >= 2:
            consistent_rows = 0
            for line in lines[1:]:
                if len(line.split(sep)) >= header_cells - 1 and len(line.split(sep)) <= header_cells + 1:
                    consistent_rows += 1
            return consistent_rows >= len(lines) / 2 and consistent_rows > 0
    return False


# Function to fetch data from the database
def fetch_data(table_name):
    connection = get_db_connection()
    if connection is None:
        return None, "Failed to connect to the database."

    try:
        with connection.cursor() as cursor:
            query = f"SELECT * FROM {table_name}"
            cursor.execute(query)
            results = cursor.fetchall()
            field_names = [i[0] for i in cursor.description]
            return results, field_names
    except pymysql.MySQLError as e:
        error_message = f"Error fetching data from {table_name}: {e}"
        print(error_message)
        flash(error_message, 'error')
        return None, error_message
    finally:
        connection.close()

def fetch_dropdown_data(table_name):
    """Fetches data for dropdown menus."""
    connection = get_db_connection()
    if connection is None:
        return []

    try:
        with connection.cursor() as cursor:
            query = f"SELECT {table_name.lower()}ID  FROM {table_name.lower()}"
            cursor.execute(query)
            results = cursor.fetchall()
            return results
    except pymysql.MySQLError as e:
        error_message = f"Error fetching dropdown data from {table_name}: {e}"
        print(e)
        flash(error_message, 'error')
        return []
    finally:
        connection.close()

def _fetch_recipe_details_from_db(recipe_id):
    """
    Fetches recipe details from the database given a recipe ID.

    Args:
        recipe_id (str): The ID of the recipe to fetch.

    Returns:
        dict: A dictionary containing the recipe details, or None if not found.
    """
    connection = get_db_connection()
    if not connection:
        return None

    try:
        with connection.cursor(pymysql.cursors.DictCursor) as cursor:
            query = """
                SELECT
                    r.recipeID,
                    ro.description AS role_description,
                    c.description AS command_description,
                    t.description AS topic_description,
                    con.description AS context_description,
                    ta.description AS targetaudience_description,
                    oft.description AS outputformat_description,
                    os.description AS outputstructure_description,
                    oq.description AS outputquality_description,
                    inf.description AS inputfilter_description,
                    ouf.description AS outputfilter_description,
                    conf.description AS contextfilter_description
                FROM recipe r
                LEFT JOIN role ro ON r.roleID = ro.roleID
                LEFT JOIN command c ON r.commandID = c.commandID
                LEFT JOIN topic t ON r.topicID = t.topicID
                LEFT JOIN context con ON r.contextID = con.contextID
                LEFT JOIN targetaudience ta ON r.targetaudienceID = ta.targetaudienceID
                LEFT JOIN outputformat oft ON r.outputformatID = oft.outputformatID
                LEFT JOIN outputstructure os ON r.outputstructureID = os.outputstructureID
                LEFT JOIN outputquality oq ON r.outputqualityID = oq.outputqualityID
                LEFT JOIN inputfilter inf ON r.inputfilterID = inf.inputfilterID
                LEFT JOIN outputfilter ouf ON r.outputfilterID = ouf.outputfilterID
                LEFT JOIN contextfilter conf ON r.contextfilterID = conf.contextfilterID
                WHERE r.recipeID = %s
            """
            cursor.execute(query, (recipe_id,))
            recipe_details = cursor.fetchone()
            return recipe_details
    except pymysql.MySQLError as e:
        print(f"Database error fetching recipe details: {e}")
        return None
    finally:
        connection.close()



# Function to update data in the database
def update_data(table_name, primary_key_value, data):
    connection = get_db_connection()
    if connection is None:
        return False, "Failed to connect to the database."

    try:
        with connection.cursor() as cursor:
            set_clause = ", ".join(f"`{field}` = %s" for field in data.keys())
            query = f"UPDATE {table_name} SET {set_clause} WHERE {table_name.lower()}ID = %s"
            values = list(data.values()) + [primary_key_value]
            cursor.execute(query, values)
            connection.commit()
        return True, None
    except pymysql.MySQLError as e:
        error_message = f"Error updating data in {table_name}: {e}"
        print(error_message)
        flash(error_message, 'error')
        return False, error_message
    finally:
        connection.close()

# Function to delete data from the database
def delete_data(table_name, primary_key_value):
    connection = get_db_connection()
    if connection is None:
        return False, "Failed to connect to the database."

    try:
        with connection.cursor() as cursor:
            query = f"DELETE FROM {table_name} WHERE {table_name.lower()}ID = %s"
            cursor.execute(query, (primary_key_value,))
            connection.commit()
        return True, None
    except pymysql.MySQLError as e:
        error_message = f"Error deleting data from {table_name}: {e}"
        print(error_message)
        flash(error_message, 'error')
        return False, error_message
    finally:
        connection.close()

# Function to add new data to the database
def add_data(table_name, data):
    connection = get_db_connection()
    if connection is None:
        return False, "Failed to connect to the database."
    try:
        with connection.cursor() as cursor:
            columns = ", ".join(f"`{field}`" for field in data.keys())
            placeholders = ", ".join("%s" for _ in data.values())
            query = f"INSERT INTO {table_name} ({columns}) VALUES ({placeholders})"
            values = list(data.values())
            cursor.execute(query, values)
            connection.commit()
        return True, None
    except pymysql.MySQLError as e:
        error_message = f"Error adding data to {table_name}: {e}"
        print(error_message)
        flash(error_message, 'error')
        return False, error_message
    finally:
        connection.close()

@app.route('/')
@login_required # Protect the index page
def index():
    is_admin = session.get('role') == 'admin'
    is_bot_user = session.get('role') == 'bot' # Changed 'bot_user' to 'bot'

    # If it's a bot_user, redirect directly to the chatbot page
    if is_bot_user:
        return redirect(url_for('chatbot'))

    print(config_table)
    # Pass tables to index.html for the menu
    return render_template('index.html', tables=config_init(), is_admin=is_admin, is_bot_user=is_bot_user)

@app.route('/table/<table_name>')
@admin_required # Only admin can view tables
def show_table(table_name):
    conn = get_db_connection()
    if conn is None:
        return render_template('error.html', message="Failed to connect to the database.")

    cursor = conn.cursor()
    cursor.execute("SHOW TABLES")
    available_tables = [row[0] for row in cursor.fetchall()]
    if table_name not in available_tables:
        cursor.close()
        conn.close()
        flash(f"Table '{table_name}' does not exist.", 'error')
        return render_template('error.html', message=f"Table '{table_name}' does not exist.")

    try:
        cursor.execute(f"SELECT * FROM {table_name}")
        rows = cursor.fetchall()
        columns = [desc[0] for desc in cursor.description]
    except pymysql.MySQLError as e:
        error_message = f"Error fetching data from {table_name}: {e}"
        print(error_message)
        flash(error_message, 'error')
        cursor.close()
        conn.close()
        return render_template('error.html', message=error_message)
    finally:
        cursor.close()
        conn.close()

    # Pass tables to data.html (which extends index.html)
    return render_template('data.html', table_name=table_name, field_names=columns, data=rows, tables=config_init())

@app.route('/modify', methods=['POST'])
@admin_required # Only admin can modify
def modify():
    selected_ids = request.form.getlist('selected_ids')
    table_name = request.form.get('table_name')
    if not selected_ids:
        flash('Please select at least one item to modify.', 'warning')
        return redirect(url_for('show_table', table_name=table_name))

    if len(selected_ids) > 1:
        flash('You can only modify one item at a time.', 'warning')
        return redirect(url_for('show_table', table_name=table_name))

    primary_key_value = selected_ids[0]
    connection = get_db_connection()
    if connection is None:
        return render_template('error.html', message="Failed to connect to the database.")

    try:
        with connection.cursor() as cursor:
            query = f"SELECT * FROM {table_name} WHERE {table_name.lower()}ID = %s"
            cursor.execute(query, (primary_key_value,))
            result = cursor.fetchone()
            field_names = [i[0] for i in cursor.description]
            if result:
                data = dict(zip(field_names, result))
                # Pass tables to modify.html (which extends index.html)
                return render_template('modify.html', table_name=table_name, data=data, tables=config_init())
            else:
                flash(f"Item with ID {primary_key_value} not found in {table_name}.", 'error')
                return redirect(url_for('show_table', table_name=table_name))
    except pymysql.MySQLError as e:
        error_message = f"Error fetching data for modification: {e}"
        print(error_message)
        flash(error_message, 'error')
        return render_template('error.html', message=error_message)
    finally:
        connection.close()

@app.route('/confirm_modify', methods=['POST'])
@admin_required # Only admin can confirm modify
def confirm_modify():
    table_name = request.form.get('table_name')
    primary_key_value = request.form.get(f'{table_name.lower()}ID')
    data = {}
    for key, value in request.form.items():
        if key not in ('table_name', f'{table_name.lower()}ID'):
            data[key] = value
        table_name1 = table_name.capitalize()
        primary_key_column = f"{table_name}ID"
        primary_key_value = request.form.get(primary_key_column)
    print(f"Table Name: {table_name}")
    print(f"Primary Key Value: {primary_key_value}")
    print(f"Data: {data}")

    success, error_message = update_data(table_name, primary_key_value, data)
    if success:
        return redirect(url_for('show_table', table_name=table_name))
    else:
        # If modify fails, render error page and pass tables for menu
        return render_template('error.html', message=error_message, tables=config_init())

@app.route('/delete', methods=['POST'])
@admin_required # Only admin can delete
def delete():
    selected_ids = request.form.getlist('selected_ids')
    table_name = request.form.get('table_name')
    if not selected_ids:
        flash('Please select at least one item to delete.', 'warning')
        return redirect(url_for('show_table', table_name=table_name))

    if len(selected_ids) > 1:
        flash('You can only delete one item at a time.', 'warning')
        return redirect(url_for('show_table', table_name=table_name))

    primary_key_value = selected_ids[0]
    success, error_message = delete_data(table_name, primary_key_value)
    if success:
        flash(f'{table_name} data deleted successfully!', 'success')
        return redirect(url_for('show_table', table_name=table_name))
    else:
        # If delete fails, render error page and pass tables for menu
        return render_template('error.html', message=error_message, tables=config_init())
    
@app.route('/add', methods=['GET', 'POST'])
@admin_required # Only admin can add
def add():
    table_name = request.args.get('table_name')
    if request.method == 'POST':
        data = {}
        for key, value in request.form.items():
            if key != 'table_name':
                data[key] = value
        success, error_message = add_data(table_name, data)
        if success:
            #flash(f'{table_name} data added successfully!', 'success')
            return redirect(url_for('show_table', table_name=table_name))
        else:
            # If add fails, render error page and pass tables for menu
            return render_template('error.html', message=error_message, tables=config_init())
    else:
        connection = get_db_connection()
        if connection is None:
            return render_template('error.html', message="Failed to connect to the database.")
        try:
            with connection.cursor() as cursor:
                query = f"DESCRIBE {table_name}"
                cursor.execute(query)
                columns = cursor.fetchall()
                field_names = [column[0] for column in columns]
                # Pass tables to add.html (which extends index.html)
                return render_template('add.html', table_name=table_name, field_names=field_names, tables=config_init())
        except pymysql.MySQLError as e:
            error_message = f"Error describing table {table_name}: {e}"
            print(error_message)
            flash(error_message, 'error')
            return render_template('error.html', message=error_message)
        finally:
            connection.close()

# Define the allowed file extensions
ALLOWED_EXTENSIONS = {'pdf', 'txt', 'docx', 'xlsx', 'png', 'jpg', 'jpeg'}

def allowed_file(filename):
    """Checks if the file extension is allowed."""
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

# Function to handle file uploads
def handle_file_upload(file):
    """
    Handles file upload and saves it to a designated directory.
    Returns the filename and filepath if successful, None otherwise.
    """
    if file and allowed_file(file.filename):
        filename = secure_filename(file.filename)
        upload_dir = 'uploads'
        if not os.path.exists(upload_dir):
            os.makedirs(upload_dir)
        filepath = os.path.join(upload_dir, filename)
        try:
            file.save(filepath)
            return filename, filepath
        except Exception as e:
            print(f"Error saving file: {e}")
            flash(f"Error saving file: {e}", 'error')
            return None, None
    elif file:
        flash('Invalid file type. Please upload a pdf, txt, docx, xlsx, png, or jpg file.', 'error')
        return None, None
    return None, None

def read_file_content(filepath):
    """
    Reads the content of a file based on its type.
    For images, returns base64 encoded string. For others, returns text.

    Args:
        filepath (str): The path to the file.

    Returns:
        tuple: (content_type, content_data) where content_type is 'text' or 'image_base64'.
               content_data is the extracted text or base64 string.
               Returns (None, None) on error or if image processing fails for an image file.
    """
    print("Entered read_file_content")
    try:
        file_extension = filepath.rsplit('.', 1)[1].lower()
        if file_extension in {'png', 'jpg', 'jpeg'}:
            if Image: # Check if PIL is imported and available
                with open(filepath, 'rb') as image_file:
                    encoded_string = base64.b64encode(image_file.read()).decode('utf-8')
                    # Data URI format: data:image/<extension>;base64,<base64_string>
                    return 'image_base64', f"data:image/{file_extension};base64,{encoded_string}"
            else:
                print("Pillow (PIL) not installed, cannot process images directly.")
                flash("Image processing not supported. Please install Pillow.", 'error')
                return None, None # Return None, None if image processing is not possible
        elif file_extension == 'pdf':
            if PyPDF2:
                with open(filepath, 'rb') as file:
                    reader = PyPDF2.PdfReader(file)
                    text = ""
                    for page in reader.pages:
                        text += page.extract_text() or ""
                    text = redact_pii(text)  # Redact PII from PDF text
                    return 'text', text
            else:
                print("PyPDF2 not installed, cannot process PDFs.")
                flash("PDF processing not supported. Please install PyPDF2.", 'error')
                return None, None
        elif file_extension == 'txt':
            with open(filepath, 'r', encoding='utf-8', errors='ignore') as file:
                text = file.read()
                text = redact_pii(text)  # Redact PII from TXT text
                return 'text', text
        elif file_extension == 'docx':
            if docx:
                print("Within read_file_content :: if docx")
                doc = docx.Document(filepath)
                text = ""
                for paragraph in doc.paragraphs:
                    text += paragraph.text + "\n"
                text = redact_pii(text)  # Redact PII from DOCX text
                return 'text', text
            else:
                print("python-docx not installed, cannot process DOCX files.")
                flash("DOCX processing not supported. Please install python-docx.", 'error')
                return None, None
        elif file_extension == 'xlsx':
            if openpyxl:
                workbook = openpyxl.load_workbook(filepath)
                sheet = workbook.active
                text = ""
                for row in sheet.iter_rows():
                    row_text = ""
                    for cell in row:
                        row_text += str(cell.value) + "\t"
                    text += row_text.strip() + "\n"
                text = redact_pii(text)  # Redact PII from XLSX text
                return 'text', text
            else:
                print("openpyxl not installed, cannot process XLSX files.")
                flash("XLSX processing not supported. Please install openpyxl.", 'error')
                return None, None
        else:
            flash(f"Unsupported file format: .{file_extension}", 'error')
            return None, None  # Explicitly return None, None for unsupported formats
    except Exception as e:
        print(f"Error reading file content from {filepath}: {e}")
        flash(f"Error processing file: {e}", 'error')
        return None, None # Return None, None on any error during file reading
    

def get_openai_response(prompt_text, image_url=None):
    """
    Sends a prompt to Azure OpenAI and returns the response.
    Can include an image URL for vision-capable models.
    """
    openai.api_type = "azure"
    openai.api_key = azure_openai_config["api_key"]
    openai.api_base = azure_openai_config["endpoint"]
    openai.api_version = azure_openai_config["api_version"]
    deployment_name = azure_openai_config["deployment_name"] # This is the 'engine' or 'deployment_id'

    messages_content = []

    # Add text content
    messages_content.append({"type": "text", "text": prompt_text})

    # Add image content if provided
    if image_url:
        messages_content.append({
            "type": "image_url",
            "image_url": {"url": image_url} # This is the Base64 data URI
        })

    try:
        response = openai.ChatCompletion.create(
            engine=deployment_name, # Use 'engine' for Azure OpenAI for deployment name
            messages=[
                {"role": "user", "content": messages_content}
            ],
            max_tokens=15000 # You might want to adjust this based on expected response length
        )
        #print(response)
        return response.choices[0].message.content.strip()
    except Exception as e:
        print(f"Error with Azure OpenAI: {e}")
        flash(f"Error with Azure OpenAI: {e}", 'error')
        return "Failed to get response from OpenAI."

# New route for the chatbot screen

@app.route('/chatbot')
@chatbot_access_required  # Protect chatbot access
def chatbot():
    user_id = session.get('user_id')
    user_role = session.get('role')  # Get the user's role
    connection = get_db_connection()

    if connection:
        try:
            with connection.cursor() as cursor:
                if user_role == 'admin':
                    # If the user is an admin, fetch all projects
                    cursor.execute("SELECT DISTINCT pi.projectID FROM userprompts up JOIN promptinstance pi ON up.promptinstanceID = pi.promptinstanceID")
                    projects = cursor.fetchall()
                else:
                    # For non-admin users, fetch projects based on userprompts join
                    cursor.execute(
                        """
                        SELECT DISTINCT pi.projectID
                        FROM userprompts up
                        JOIN promptinstance pi ON up.promptinstanceID = pi.promptinstanceID
                        WHERE up.userID = %s
                        """,
                        (user_id,),
                    )
                    projects = cursor.fetchall()

                if not projects:
                    # User not found or no associated projects
                    flash("No Projects associated with this user.", 'warning')
                    projects = []

                usecases = fetch_dropdown_data('Usecase')
                recipes = fetch_dropdown_data('Recipe')

                return render_template(
                    'chatbot.html',
                    projects=projects,
                    usecases=usecases,
                    recipes=recipes,
                    tables=config_init(),
                )

        except pymysql.MySQLError as e:
            print(f"Database error: {e}")
            flash(f"Database error: {e}", 'error')
            return render_template('chatbot.html', projects=[], usecases=[], recipes=[], tables=config_init())
        finally:
            connection.close()
    else:
        flash("Failed to connect to the database.", 'error')
        return render_template('chatbot.html', projects=[], usecases=[], recipes=[], tables=config_init())
    


@app.route('/chatbot/get_usecases_for_project/<project_id>')  
@chatbot_access_required
def get_usecases_for_project(project_id):
        user_id = session.get('user_id')
        user_role = session.get('role')  # Get the user's role
        connection = get_db_connection()
        if connection is None:
            return jsonify({'error': "Failed to connect to the database."}), 500

        try:
            with connection.cursor() as cursor:
                if user_role == 'admin':
                        # If the user is an admin, fetch all usecases
                        cursor.execute(
                            """
                            SELECT DISTINCT u.usecaseID
                            FROM usecase u
                            JOIN promptinstance pi ON u.usecaseID = pi.usecaseID
                            WHERE pi.projectID = %s
                            """,
                            (project_id,),
                        )
                        usecases = [row[0] for row in cursor.fetchall()]

                else:
                    cursor.execute(
                            """
                            SELECT DISTINCT u.usecaseID
                            FROM userprompts up
                            JOIN promptinstance pi ON up.promptinstanceID = pi.promptinstanceID
                            JOIN usecase u ON pi.usecaseID = u.usecaseID
                            WHERE up.userID = %s AND pi.projectID = %s

                            """,
                            (user_id, project_id),
                        )
                    usecases = [row[0] for row in cursor.fetchall()]
                return jsonify({'usecases': usecases})
        except pymysql.MySQLError as e:
            error_message = f"Database error fetching usecases: {e}"
            print(error_message)
            return jsonify({'error': error_message}), 500
        finally:
            connection.close()

@app.route('/chatbot/get_recipes_for_project_usecase/<project_id>/<usecase_id>')  # Removed <int:> for project_id
@chatbot_access_required
def get_recipes_for_project_usecase(project_id, usecase_id):
    user_id = session.get('user_id') # Get the current user's ID from the session
    user_role = session.get('role') # Get the user's role

    connection = get_db_connection()
    if connection is None:
        return jsonify({'error': "Failed to connect to the database."}), 500

    try:
        with connection.cursor() as cursor:
            if user_role == 'admin':
                # Admins can see all recipes for the selected project and usecase
                cursor.execute("""
                    SELECT DISTINCT r.recipeID
                    FROM recipe r
                    JOIN promptinstance pi ON r.recipeID = pi.recipeID
                    WHERE pi.projectID = %s AND pi.usecaseID = %s
                """, (project_id, usecase_id))
            else:
                # For non-admin users, filter recipes based on userprompts table
                cursor.execute("""
                    SELECT DISTINCT r.recipeID
                    FROM recipe r
                    JOIN promptinstance pi ON r.recipeID = pi.recipeID
                    WHERE pi.projectID = %s
                      AND pi.usecaseID = %s
                      AND pi.promptinstanceID IN (
                          SELECT up.promptinstanceID
                          FROM userprompts up
                          WHERE up.userID = %s
                      )
                """, (project_id, usecase_id, user_id))

            recipes = [row[0] for row in cursor.fetchall()]
            return jsonify({'recipes': recipes})
    except pymysql.MySQLError as e:
        error_message = f"Database error fetching recipes: {e}"
        print(error_message)
        return jsonify({'error': error_message}), 500
    finally:
        connection.close()



@app.route('/chatbot/get_recipe_details/<recipe_id>')
@chatbot_access_required
def get_recipe_details(recipe_id):
    recipe_details = _fetch_recipe_details_from_db(recipe_id)
    if recipe_details:
        recipe_details_text = "\n".join([
            f"Role : {recipe_details.get('role_description', '')}",
            f"Command : {recipe_details.get('command_description', '')}",
            f"Topic : {recipe_details.get('topic_description', '')}",
            f"Context : {recipe_details.get('context_description', '')}",
            f"Target Audience : {recipe_details.get('targetaudience_description', '')}",
            f"Output Format : {recipe_details.get('outputformat_description', '')}",
            f"Output Structure : {recipe_details.get('outputstructure_description', '')}",
            f"Output Quality : {recipe_details.get('outputquality_description', '')}",
            f"Input Filter : {recipe_details.get('inputfilter_description', '')}",
            f"Output Filter : {recipe_details.get('outputfilter_description', '')}",
            f"Context Filter : {recipe_details.get('contextfilter_description', '')}",
        ])
        return jsonify({'recipe_details': recipe_details_text})
    else:
        return jsonify({'recipe_details': 'Recipe details not found.'})



# Route to handle chatbot requests

@app.route('/chatbot/get_response', methods=['POST'])
@chatbot_access_required
def get_response():
    print("Starting get_response")
    print(request)
    print(request.form)
    project_id = request.form.get('project_id')
    print(f"Project ID {project_id}")
    usecase_id = request.form.get('usecase_id')
    recipe_id = request.form.get('recipe_id')
    message = request.form.get('message')
    print(f"recipe ID {recipe_id}")
    file = request.files.get('file')
    print("Inside get response")
    print(file)
    uploaded_filepath = None
    file_content_type = None
    file_content_data = None

    # Initialize variables for generated file
    generated_file_data = None
    generated_file_type = None
    generated_file_name = None
    mime_type = 'text/plain' # Default MIME type

    if recipe_id == "Empty":
        flash("The 'Empty' recipe was selected. Only your message will be sent to the LLM.", 'info')
        prompt_text = f"User message: {message}"
        openai_response = get_openai_response(prompt_text)
        # Default to text download for "Empty" recipe response
        generated_file_data = None
        generated_file_type = None
        generated_file_name = None
        mime_type = 'text/plain'
        return jsonify({
            'response': openai_response,
            'file_data': generated_file_data,
            'file_type': generated_file_type,
            'file_name': generated_file_name,
            'mime_type': mime_type
        })
    
    # Flag to determine if LLM call should proceed
    proceed_with_llm = True 

    if file:
        if allowed_file(file.filename):
            print("Inside :: allowed_file(file.filename)")
            uploaded_filename = secure_filename(file.filename)
            upload_dir = app.config['UPLOAD_FOLDER'] # Use the configured UPLOAD_FOLDER
            if not os.path.exists(upload_dir):
                os.makedirs(upload_dir) # Ensure directory exists
            uploaded_filepath = os.path.join(upload_dir, uploaded_filename)

            try:
                file.save(uploaded_filepath)
                print("File.save successful")
                file_content_type, file_content_data = read_file_content(uploaded_filepath)
                if file_content_type is None: # If read_file_content returned None, None due to error/unsupported
                    proceed_with_llm = False
                    flash("Failed to process uploaded file content.", 'error')
            except Exception as e:
                print(f"Error saving or reading file: {e}")
                flash(f"Error saving or reading file: {e}", 'error')
                proceed_with_llm = False
        else:
            flash(f"Unsupported file format. Allowed formats are: {', '.join(ALLOWED_EXTENSIONS)}", 'danger')
            proceed_with_llm = False # Do not proceed with LLM if file format is not allowed

    recipe_details = _fetch_recipe_details_from_db(recipe_id)
    recipe_details_prompt_part = ""

    if recipe_details:
        # Extract descriptions, handling None and "Empty" values
        role_desc = recipe_details.get('role_description')
        role_desc = role_desc if role_desc and role_desc.strip().lower() != "empty" else None
        command_desc = recipe_details.get('command_description')
        command_desc = command_desc if command_desc and command_desc.strip().lower() != "empty" else None
        topic_desc = recipe_details.get('topic_description')
        topic_desc = topic_desc if topic_desc and topic_desc.strip().lower() != "empty" else None
        context_desc = recipe_details.get('context_description')
        context_desc = context_desc if context_desc and context_desc.strip().lower() != "empty" else None
        targetaudience_desc = recipe_details.get('targetaudience_description')
        targetaudience_desc = targetaudience_desc if targetaudience_desc and targetaudience_desc.strip().lower() != "empty" else None
        outputformat_desc = recipe_details.get('outputformat_description')
        outputformat_desc = outputformat_desc if outputformat_desc and outputformat_desc.strip().lower() != "empty" else None
        outputstructure_desc = recipe_details.get('outputstructure_description')
        outputstructure_desc = outputstructure_desc if outputstructure_desc and outputstructure_desc.strip().lower() != "empty" else None
        outputquality_desc = recipe_details.get('outputquality_description')
        outputquality_desc = outputquality_desc if outputquality_desc and outputquality_desc.strip().lower() != "empty" else None
        
        # Format the prompt string for the LLM, only including non-None components
        recipe_details_prompt_part = ""
        if role_desc:
            recipe_details_prompt_part += f" {role_desc} \n"
        if command_desc:
            recipe_details_prompt_part += f" {command_desc} \n"
        if topic_desc:
            recipe_details_prompt_part += f" {topic_desc} \n"
        if context_desc:
            recipe_details_prompt_part += f" {context_desc} "
        if targetaudience_desc:
            recipe_details_prompt_part += f" {targetaudience_desc} "
        if outputformat_desc:
            recipe_details_prompt_part += f" {outputformat_desc} "
        if outputstructure_desc:
            recipe_details_prompt_part += f" {outputstructure_desc}\n"
        if outputquality_desc:
            recipe_details_prompt_part += f" {outputquality_desc} \n"

        
    else:
        recipe_details_prompt_part = f"Recipe ID: {recipe_id}. "
        flash(f"Recipe details not found for Recipe ID: {recipe_id}", 'warning')

    # Construct the full prompt text for the LLM
    # PII Redaction (assuming redact_pii function is defined elsewhere in app.py)
    message_to_llm = redact_pii(message)
    prompt_text = f"{recipe_details_prompt_part} {message_to_llm}"
    #print(f" Prompt Text that is sent to LLM: {prompt_text}")

    openai_response = "No response generated due to file processing issues or unsupported format."
    if proceed_with_llm: # Only call LLM if file processing was successful or no file was uploaded
        # Pass image_url if the file content is an image (Base64 data URI) AND it's valid
        image_url_to_send = file_content_data if file_content_type == 'image_base64' and file_content_data else None

        # If it's a text file, append its content to the prompt_text
        if file_content_type == 'text' and file_content_data:
            prompt_text += f"\nFile content: {file_content_data}"

        openai_response = get_openai_response(prompt_text, image_url=image_url_to_send)

        # --- Dynamic File Type Determination and Data Preparation ---
        # This is where you decide what type of file to offer for download
        if "feature" in openai_response.lower() and ("scenario:" in openai_response.lower() or "feature:" in openai_response.lower()):
            generated_file_data = openai_response
            generated_file_type = "feature"
            generated_file_name = "llm_output.feature"
            mime_type = 'text/plain'
            openai_response += "\n\nGenerated a feature file. Click 'Download'."
        elif is_likely_excel(openai_response):
            # Convert TSV-like data to Excel
            if openpyxl:
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "LLM Data"
                lines = openai_response.strip().split('\n')
                separators = ['|', '\t']
                used_separator = None
                for sep in separators:
                    if len(lines[0].split(sep)) >= 2:
                        used_separator = sep
                        break

                if used_separator:
                    for line in lines:
                        ws.append([cell.strip() for cell in line.split(used_separator)])
                    output = io.BytesIO()
                    wb.save(output)
                    output.seek(0)
                    generated_file_data = base64.b64encode(output.read()).decode('utf-8')
                    generated_file_type = "xlsx"
                    generated_file_name = "llm_data.xlsx"
                    mime_type = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                    openai_response += "\n\nGenerated Excel data. Click 'Download'."
                else:
                    # Fallback to text if no clear separator is found
                    generated_file_data = openai_response
                    generated_file_type = "txt"
                    generated_file_name = "llm_data.txt"
                    mime_type = 'text/plain'
                    openai_response += "\n\nGenerated data (as text due to format detection issue). Click 'Download'."
            else:
                generated_file_data = openai_response
                generated_file_type = "txt"
                generated_file_name = "llm_data.txt"
                mime_type = 'text/plain'
                openai_response += "\n\nCould have generated Excel, but openpyxl is not installed. Download as text."
        elif "excel" in openai_response.lower() and openpyxl:
            # Handle explicit "excel" request (similar to TSV handling now)
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "LLM Response"
            lines = openai_response.split('\n')
            for line in lines:
                if '\t' in line:
                    ws.append(line.split('\t'))
                elif '|' in line:
                    ws.append([cell.strip() for cell in line.strip('|').split('|')])
                else:
                    ws.append([line])
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            generated_file_data = base64.b64encode(output.read()).decode('utf-8')
            generated_file_type = "xlsx"
            generated_file_name = "llm_report.xlsx"
            mime_type = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            openai_response += "\n\nGenerated an Excel file. Click 'Download'."
        else:
            generated_file_data = openai_response
            generated_file_type = "txt"
            generated_file_name = "chatbot_response.txt"
            mime_type = 'text/plain'

    else:
        openai_response = "LLM call skipped due to file processing errors or unsupported file type."

    # Clean up the uploaded file after processing
    if uploaded_filepath and os.path.exists(uploaded_filepath):
        try:
            os.remove(uploaded_filepath)
            print(f"Cleaned up uploaded file: {uploaded_filepath}")
        except Exception as e:
            print(f"Error cleaning up file {uploaded_filepath}: {e}")


    return jsonify({
        'response': openai_response,
        'file_data': generated_file_data,
        'file_type': generated_file_type,
        'file_name': generated_file_name,
        'mime_type': mime_type
    })


# Login route
@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']

        connection = get_db_connection()
        if connection is None:
            flash("Database connection failed for login.", 'danger')
            return render_template('login.html')

        try:
            with connection.cursor(pymysql.cursors.DictCursor) as cursor:
                query = "SELECT UserID, Type, username, password FROM user WHERE username = %s AND password = %s"
                cursor.execute(query, (username, password))
                user = cursor.fetchone()

                if user:
                    session['logged_in'] = True
                    session['username'] = user['username']
                    session['role'] = user['Type'].lower()
                    session['user_id'] = user['UserID']

                    # Fetch all projectIDs for the user (CORRECTED QUERY)
                    cursor.execute("SELECT projectID FROM user WHERE username = %s", (username,))
                    project_rows = cursor.fetchall()
                    session['project_ids'] = [row['projectID'] for row in project_rows]

                    flash(f'Logged in as {user["username"]} ({user["Type"]})', 'success')

                    if session['role'] == 'bot':
                        return redirect(url_for('chatbot'))
                    else:
                        return redirect(url_for('index'))
                else:
                    flash('Invalid credentials', 'danger')
        except pymysql.MySQLError as e:
            error_message = f"Database error during login: {e}"
            print(error_message)
            flash(error_message, 'danger')
        finally:
            connection.close()
    return render_template('login.html')


# Logout route
@app.route('/logout')
@login_required
def logout():
    session.pop('logged_in', None)
    session.pop('username', None)
    session.pop('role', None)
    session.pop('user_id', None) # Clear user_id
    session.pop('project_id', None) # Clear project_id
    session.clear()
    flash('You have been logged out.', 'info')
    return redirect(url_for('login'))


if __name__ == "__main__":
    config_init()
    #app.run(debug=True)
    app.run(host="0.0.0.0", port=8899, debug=True)
